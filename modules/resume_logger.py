'''
Resume Logger Module
Appends one row to an Excel workbook each time a custom resume is used.

Columns logged:
  Date Applied | Company | Job Title | Resume File Path | Projects Selected | Job Description (Snippet)

The Excel file is created automatically on first run.
'''

import os
from datetime import datetime
from modules.helpers import print_lg

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False
    print_lg(
        "ResumeLogger: 'openpyxl' not installed — Excel logging disabled.\n"
        "  → Install it with:  pip install openpyxl"
    )


_HEADERS = [
    "Date Applied",
    "Company",
    "Job Title",
    "Resume File Path",
    "Projects Selected",
    "Job Description (Snippet)",
]

_HEADER_COLOR  = "4472C4"   # blue header fill
_ALT_ROW_COLOR = "DCE6F1"   # light blue alternating row fill
_COL_WIDTHS    = [22, 28, 32, 60, 60, 60]


def log_resume_application(
    company: str,
    job_title: str,
    resume_path: str,
    selected_projects: list,
    job_description: str = "",
    log_path: str = "all excels/resume_log.xlsx"
) -> None:
    '''
    Appends a row to the resume log Excel file.
    Creates the file with a styled header row if it does not exist yet.
    '''
    if not _OPENPYXL_OK:
        return

    try:
        os.makedirs(os.path.dirname(log_path) if os.path.dirname(log_path) else ".", exist_ok=True)

        if os.path.exists(log_path):
            wb = openpyxl.load_workbook(log_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resume Log"
            _write_header(ws)

        row_idx = ws.max_row + 1

        row_data = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            company or "Unknown",
            job_title or "Unknown",
            os.path.abspath(resume_path) if resume_path else "Default",
            ", ".join(selected_projects) if selected_projects else "N/A",
            (job_description or "")[:400],
        ]

        # Apply alternating row fill
        fill = PatternFill(
            start_color=_ALT_ROW_COLOR,
            end_color=_ALT_ROW_COLOR,
            fill_type="solid"
        ) if row_idx % 2 == 0 else None

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
            if fill:
                cell.fill = fill

        # Refresh column widths if they shrank
        for col_idx, width in enumerate(_COL_WIDTHS, start=1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            current = ws.column_dimensions[letter].width
            ws.column_dimensions[letter].width = max(current or 0, width)

        wb.save(log_path)
        print_lg(f"ResumeLogger: Logged → {log_path}  (row {row_idx})")

    except Exception as e:
        print_lg(f"ResumeLogger: Failed to write log — {e}")


def _write_header(ws) -> None:
    '''Writes a styled header row to a fresh worksheet.'''
    header_fill   = PatternFill(start_color=_HEADER_COLOR, end_color=_HEADER_COLOR, fill_type="solid")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
